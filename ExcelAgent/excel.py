import os
import openpyxl
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openai import AzureOpenAI
import json
from datetime import datetime, timedelta
from sharepoint import download_excel, upload_excel
from typing import Optional
import uuid
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

app = FastAPI()
templates = Jinja2Templates(directory="templates")

# Initialize Azure OpenAI client
client = AzureOpenAI(
    api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    api_version=os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview"),
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT")
)


EXCEL_FILE = "contracts.xlsx"
SHAREPOINT_SYNC_ENABLED = False

HEADERS = [
    "Sr. No.",
    "Description of the Contract",
    "Name of the first Party",
    "Name of the second Party",
    "Date of Request",
    "Purpose",
    "Agreement Commencement date",
    "Duration of the Agreement",
    "Department Responsibility",
    "Internal Status",
    "Signed Copy RECEIVED on IVALUA (Y/N)",
    "Uploaded on IVALUA"
]

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)

# Add sheets mode state
sheets_mode = False

class InputRequest(BaseModel):
    input: str
    overwrite: Optional[bool] = False

class SyncRequest(BaseModel):
    action: str = "both"

class SheetsModeRequest(BaseModel):
    enabled: bool

def sync_from_sharepoint():
    if not SHAREPOINT_SYNC_ENABLED:
        return True, "SharePoint sync is disabled"
    
    try:
        print("🔄 Syncing latest file from SharePoint...")
        download_excel(EXCEL_FILE)
        print(f"✅ Successfully downloaded latest file to {EXCEL_FILE}")
        return True, "Successfully synced from SharePoint"
    except Exception as e:
        print(f"❌ Failed to sync from SharePoint: {e}")
        return False, f"SharePoint sync failed: {str(e)}"

def sync_to_sharepoint():
    if not SHAREPOINT_SYNC_ENABLED:
        return True, "SharePoint sync is disabled"
    
    try:
        print("📤 Uploading changes to SharePoint...")
        success,message,url = upload_excel(EXCEL_FILE)
        print("✅ Successfully uploaded to SharePoint")
        return True, "Successfully synced to SharePoint ", url
    except Exception as e:
        print(f"❌ Failed to upload to SharePoint: {e}")
        return False, f"SharePoint upload failed: {str(e)}"

def initialize_excel():
    sync_success, sync_message = sync_from_sharepoint()
    
    if sync_success and os.path.exists(EXCEL_FILE):
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            first_row = [cell.value for cell in ws[1]]
            if first_row != HEADERS:
                print("⚠  Header mismatch detected, updating headers...")
                ws.delete_rows(1)
                ws.insert_rows(1)
                for col, header in enumerate(HEADERS, 1):
                    ws.cell(row=1, column=col, value=header)
                wb.save(EXCEL_FILE)
                sync_to_sharepoint()
            print("✅ Excel file initialized with SharePoint sync")
            return
        except Exception as e:
            print(f"⚠  Error validating synced file: {e}")
    
    if not sync_success:
        print(f"⚠  {sync_message}. Creating local file...")
    
    if not os.path.exists(EXCEL_FILE):
        print("📝 Creating new Excel file...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contracts"
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)
        
        upload_success, upload_message,url = sync_to_sharepoint()
        if not upload_success:
            print(f"⚠  {upload_message}")

def load_workbook():
    try:
        return openpyxl.load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contracts"
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)
        return wb

def get_active_sheet():
    wb = load_workbook()
    return wb, wb.active

def is_exact_duplicate(new_data, existing_row):
    for idx, header in enumerate(HEADERS):
        new_value = str(new_data.get(header, "")).strip()
        existing_value = str(existing_row[idx] if idx < len(existing_row) else "").strip()
        if new_value != existing_value:
            return False
    return True

def check_duplicate(new_data, ws):
    sr_no_duplicate = False
    full_duplicate = False
    duplicate_row = None
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
            
        if row[0] and str(row[0]).strip() == str(new_data.get("Sr. No.", "")).strip():
            sr_no_duplicate = True
            duplicate_row = row_idx
            
        if is_exact_duplicate(new_data, row):
            full_duplicate = True
            duplicate_row = row_idx
            break
    
    return {
        "sr_no_duplicate": sr_no_duplicate,
        "full_duplicate": full_duplicate,
        "row": duplicate_row
    }

def update_excel(data, update_existing=False):
    wb, ws = get_active_sheet()
    
    duplicate_info = check_duplicate(data, ws)
    
    if duplicate_info["full_duplicate"]:
        return False, f"Exact duplicate found at row {duplicate_info['row']}", None

    if duplicate_info["sr_no_duplicate"]:
        if update_existing:
            row_idx = duplicate_info["row"]
            # Clear existing row data
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx, value="")
            # Update with new data
            for col_idx, header in enumerate(HEADERS, 1):
                ws.cell(row=row_idx, column=col_idx, value=data.get(header, ""))
            print(f"✅ Updated existing row {row_idx} with new data")
        else:
            return False, f"Duplicate Sr. No. found at row {duplicate_info['row']}", None
    else:
        row = [data.get(header, "") for header in HEADERS]
        ws.append(row)
    
    try:
        wb.save(EXCEL_FILE)
    except Exception as e:
        return False, f"Failed to save local file: {str(e)}", None
    
    upload_success, upload_message, url = sync_to_sharepoint()
    if not upload_success:
        print(f"⚠  Local file updated but SharePoint sync failed: {upload_message}")
        return True, f"Success (local). SharePoint sync failed: {upload_message}", None
    
    return True, "Success (synced to SharePoint)", url

def parse_input_with_llm(user_input):
    system_prompt = f"""
You are an expert at extracting contract information from natural language. 
Return a JSON object with EXACTLY these keys: {', '.join(HEADERS)}.

Rules:
1. Use "Y" or "N" for "Signed Copy RECEIVED on IVALUA (Y/N)"
2. Format dates as YYYY-MM-DD
3. For missing values, use empty string ("")
4. Return ONLY valid JSON (no additional text)

Example Output:
{{
  "Sr. No.": "CTR-2025-001",
  "Description of the Contract": "Music Festival Agreement",
  "Name of the first Party": "SoundWave",
  "Name of the second Party": "PartyPros",
  "Date of Request": "2025-06-01",
  "Purpose": "Organize summer concert",
  "Agreement Commencement date": "2025-07-15",
  "Duration of the Agreement": "6 months",
  "Department Responsibility": "HR",
  "Internal Status": "In Progress",
  "Signed Copy RECEIVED on IVALUA (Y/N)": "Y",
  "Uploaded on IVALUA": "2025-06-08"
}}

Input: "{user_input}"
"""

    try:
        response = client.chat.completions.create(
            model=os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "gpt-4o"),
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_input}
            ],
            temperature=0.1,
            max_tokens=500
        )

        content = response.choices[0].message.content.strip()
        
        # Clean up potential markdown or code fences
        if content.startswith("```"):
            content = content.split("```")[1].strip()
        
        parsed = json.loads(content)
        
        for key, value in parsed.items():
            if isinstance(value, str):
                parsed[key] = value.strip()
                
        return parsed
    except Exception as e:
        print(f"LLM Parsing Error: {e}\nRaw Response: {content}")
        raise ValueError(f"Failed to parse contract details: {e}")

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    initialize_excel()
    return templates.TemplateResponse("excelUI.html", {"request": request})

@app.post("/api/preview")
async def preview(data: InputRequest):
    if sheets_mode:
        # Redirect to localhost:8005
        raise HTTPException(status_code=307, detail="Redirecting to sheets mode endpoint")
    try:
        parsed_data = parse_input_with_llm(data.input)
        
        _, ws = get_active_sheet()
        duplicate_info = check_duplicate(parsed_data, ws)
        
        preview_data = {header: parsed_data.get(header, "") for header in HEADERS}
        
        return {
            "success": True,
            "preview": preview_data,
            "duplicate_info": duplicate_info
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing input: {str(e)}")

@app.post("/api/submit")
async def submit(data: InputRequest):
    if sheets_mode:
        # Redirect to localhost:8005
        raise HTTPException(status_code=307, detail="Redirecting to sheets mode endpoint")
    try:
        parsed_data = parse_input_with_llm(data.input)
        
        success, message ,url= update_excel(parsed_data, data.overwrite)
        
        if not success:
            raise HTTPException(status_code=400, detail=message)
            
        return {
            "success": True,
            "message": message,
            "sr_no": parsed_data.get("Sr. No.", ""),
            "web_url":url
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing input: {str(e)}")

@app.post("/api/sync")
async def manual_sync(data: SyncRequest):
    try:
        results = {}
        
        if data.action in ["download", "both"]:
            sync_success, sync_message = sync_from_sharepoint()
            results["download"] = {
                "success": sync_success,
                "message": sync_message
            }
        
        if data.action in ["upload", "both"]:
            upload_success, upload_message = sync_to_sharepoint()
            results["upload"] = {
                "success": upload_success,
                "message": upload_message
            }
        
        overall_success = all(result["success"] for result in results.values())
        
        return {
            "success": overall_success,
            "results": results
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Sync error: {str(e)}")

@app.get("/api/status")
async def status():
    try:
        local_file_exists = os.path.exists(EXCEL_FILE)
        local_file_size = os.path.getsize(EXCEL_FILE) if local_file_exists else 0
        
        sharepoint_connected = False
        sharepoint_error = None
        try:
            from ExcelAgent.sharepoint import get_access_token
            get_access_token()
            sharepoint_connected = True
        except Exception as e:
            sharepoint_error = str(e)
        
        return {
            "local_file": {
                "exists": local_file_exists,
                "size": local_file_size,
                "path": EXCEL_FILE
            },
            "sharepoint": {
                "enabled": SHAREPOINT_SYNC_ENABLED,
                "connected": sharepoint_connected,
                "error": sharepoint_error
            }
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Status check error: {str(e)}")

@app.post("/api/sheets-mode")
async def toggle_sheets_mode(data: SheetsModeRequest):
    global sheets_mode
    sheets_mode = data.enabled
    return {
        "success": True,
        "sheets_mode": sheets_mode,
        "message": f"Sheets mode {'enabled' if sheets_mode else 'disabled'}"
    }

@app.get("/api/sheets-mode")
async def get_sheets_mode():
    return {
        "success": True,
        "sheets_mode": sheets_mode
    }

if __name__ == "__main__":
    import uvicorn
    initialize_excel()
    uvicorn.run(app, host="0.0.0.0", port=8005)